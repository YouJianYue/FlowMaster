# -*- coding: utf-8 -*-

"""
Excel导出注解和工具类

一比一复刻参考项目的 @ExcelProperty 注解功能
提供完整的Excel导出支持

@author: FlowMaster
@since: 2025/9/19
"""

from dataclasses import dataclass, field
from typing import Optional, Any, Dict, List, Union, Callable, Type
from datetime import datetime
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
import inspect
from enum import Enum


@dataclass
class ExcelProperty:
    """
    Excel属性注解

    一比一复刻参考项目的 @ExcelProperty 注解
    用于标记需要导出到Excel的字段及其配置
    """
    value: str = ""                    # 列标题
    order: int = 0                     # 列顺序
    converter: Optional[str] = None    # 转换器类名
    width: int = 20                    # 列宽度
    format_pattern: str = ""           # 格式化模式
    ignore: bool = False               # 是否忽略

    def __post_init__(self):
        """后处理：设置默认值"""
        if not self.value:
            self.value = "未命名列"


class ExcelStyle:
    """
    Excel样式配置

    对应参考项目中Excel导出的样式设置
    """

    # 标题样式
    HEADER_FONT = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

    # 数据样式
    DATA_FONT = Font(name="微软雅黑", size=10)
    DATA_ALIGNMENT = Alignment(horizontal="left", vertical="center")

    # 边框样式
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )


class ExcelConverter:
    """
    Excel转换器基类

    对应参考项目的各种Converter类
    用于在导出时转换数据格式
    """

    @staticmethod
    def convert(value: Any, *args, **kwargs) -> str:
        """
        转换值为Excel显示格式

        Args:
            value: 原始值
            *args: 额外参数
            **kwargs: 额外关键字参数

        Returns:
            转换后的字符串值
        """
        if value is None:
            return ""
        return str(value)


class ExcelListConverter(ExcelConverter):
    """
    列表转换器

    一比一复刻参考项目的 ExcelListConverter
    将列表转换为逗号分隔的字符串
    """

    @staticmethod
    def convert(value: Any, separator: str = ",", *args, **kwargs) -> str:
        """
        将列表转换为分隔字符串

        Args:
            value: 列表值
            separator: 分隔符，默认逗号

        Returns:
            转换后的字符串
        """
        if value is None:
            return ""
        if isinstance(value, (list, tuple)):
            return separator.join(str(item) for item in value)
        return str(value)


class ExcelEnumConverter(ExcelConverter):
    """
    枚举转换器

    将枚举值转换为可读的文本
    """

    @staticmethod
    def convert(value: Any, enum_class: Type[Enum] = None, *args, **kwargs) -> str:
        """
        将枚举转换为文本

        Args:
            value: 枚举值
            enum_class: 枚举类

        Returns:
            转换后的字符串
        """
        if value is None:
            return ""
        if isinstance(value, Enum):
            return value.name
        if enum_class and hasattr(enum_class, value):
            return enum_class(value).name
        return str(value)


class ExcelDateTimeConverter(ExcelConverter):
    """
    日期时间转换器

    将日期时间对象转换为指定格式的字符串
    """

    @staticmethod
    def convert(value: Any, format_pattern: str = "%Y-%m-%d %H:%M:%S", *args, **kwargs) -> str:
        """
        将日期时间转换为字符串

        Args:
            value: 日期时间值
            format_pattern: 格式化模式

        Returns:
            转换后的字符串
        """
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime(format_pattern)
        return str(value)


class ExcelBooleanConverter(ExcelConverter):
    """
    布尔值转换器

    将布尔值转换为中文
    """

    @staticmethod
    def convert(value: Any, true_text: str = "是", false_text: str = "否", *args, **kwargs) -> str:
        """
        将布尔值转换为中文

        Args:
            value: 布尔值
            true_text: 真值显示文本
            false_text: 假值显示文本

        Returns:
            转换后的字符串
        """
        if value is None:
            return ""
        if isinstance(value, bool):
            return true_text if value else false_text
        return str(value)


class ExcelExporter:
    """
    Excel导出器

    一比一复刻参考项目的Excel导出功能
    提供完整的Excel文件生成能力
    """

    def __init__(self):
        self.converters = {
            'ExcelListConverter': ExcelListConverter,
            'ExcelEnumConverter': ExcelEnumConverter,
            'ExcelDateTimeConverter': ExcelDateTimeConverter,
            'ExcelBooleanConverter': ExcelBooleanConverter,
        }

    def export(self, data: List[Any], model_class: type, filename: str = "export") -> BytesIO:
        """
        导出数据到Excel

        Args:
            data: 要导出的数据列表
            model_class: 数据模型类
            filename: 文件名

        Returns:
            Excel文件的字节流
        """
        # 获取字段配置
        fields = self._get_excel_fields(model_class)

        # 创建工作簿
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "数据导出"

        # 写入标题行
        self._write_headers(worksheet, fields)

        # 写入数据行
        self._write_data(worksheet, data, fields)

        # 应用样式
        self._apply_styles(worksheet, len(fields), len(data))

        # 保存到字节流
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        workbook.close()

        return output

    def _get_excel_fields(self, model_class: type) -> List[Dict[str, Any]]:
        """
        获取模型类的Excel字段配置

        Args:
            model_class: 模型类

        Returns:
            字段配置列表
        """
        fields = []

        # 检查是否是Pydantic模型
        if hasattr(model_class, '__fields__'):
            # Pydantic v2
            for field_name, field_info in model_class.model_fields.items():
                excel_config = self._extract_excel_config(field_info)
                if excel_config and not excel_config.ignore:
                    fields.append({
                        'name': field_name,
                        'config': excel_config,
                        'title': excel_config.value or field_name
                    })
        else:
            # 普通类，检查类属性
            for attr_name in dir(model_class):
                if not attr_name.startswith('_'):
                    attr = getattr(model_class, attr_name)
                    if hasattr(attr, 'excel_property'):
                        excel_config = attr.excel_property
                        if not excel_config.ignore:
                            fields.append({
                                'name': attr_name,
                                'config': excel_config,
                                'title': excel_config.value or attr_name
                            })

        # 按order排序
        fields.sort(key=lambda x: x['config'].order)
        return fields

    def _extract_excel_config(self, field_info) -> Optional[ExcelProperty]:
        """
        从字段信息中提取Excel配置

        Args:
            field_info: 字段信息

        Returns:
            Excel配置或None
        """
        # 检查json_schema_extra中的excel配置
        if hasattr(field_info, 'json_schema_extra') and field_info.json_schema_extra:
            excel_data = field_info.json_schema_extra.get('excel')
            if excel_data:
                return ExcelProperty(**excel_data)

        # 检查description作为默认title
        if hasattr(field_info, 'description') and field_info.description:
            return ExcelProperty(value=field_info.description)

        return None

    def _write_headers(self, worksheet, fields: List[Dict[str, Any]]):
        """
        写入表头

        Args:
            worksheet: 工作表
            fields: 字段配置列表
        """
        for col_idx, field in enumerate(fields, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.value = field['title']
            cell.font = ExcelStyle.HEADER_FONT
            cell.fill = ExcelStyle.HEADER_FILL
            cell.alignment = ExcelStyle.HEADER_ALIGNMENT
            cell.border = ExcelStyle.BORDER

            # 设置列宽
            column_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[column_letter].width = field['config'].width

    def _write_data(self, worksheet, data: List[Any], fields: List[Dict[str, Any]]):
        """
        写入数据

        Args:
            worksheet: 工作表
            data: 数据列表
            fields: 字段配置列表
        """
        for row_idx, item in enumerate(data, 2):  # 从第2行开始
            for col_idx, field in enumerate(fields, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)

                # 获取字段值
                value = self._get_field_value(item, field['name'])

                # 应用转换器
                converted_value = self._apply_converter(value, field['config'])

                cell.value = converted_value
                cell.font = ExcelStyle.DATA_FONT
                cell.alignment = ExcelStyle.DATA_ALIGNMENT
                cell.border = ExcelStyle.BORDER

    def _get_field_value(self, item: Any, field_name: str) -> Any:
        """
        获取对象字段值

        Args:
            item: 数据对象
            field_name: 字段名

        Returns:
            字段值
        """
        if hasattr(item, field_name):
            return getattr(item, field_name)
        elif isinstance(item, dict):
            return item.get(field_name)
        return None

    def _apply_converter(self, value: Any, config: ExcelProperty) -> str:
        """
        应用数据转换器

        Args:
            value: 原始值
            config: Excel配置

        Returns:
            转换后的值
        """
        if config.converter and config.converter in self.converters:
            converter_class = self.converters[config.converter]
            return converter_class.convert(value)

        # 默认转换
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(value, bool):
            return "是" if value else "否"
        if isinstance(value, (list, tuple)):
            return ",".join(str(item) for item in value)

        return str(value)

    def _apply_styles(self, worksheet, col_count: int, row_count: int):
        """
        应用Excel样式

        Args:
            worksheet: 工作表
            col_count: 列数
            row_count: 数据行数
        """
        # 冻结首行
        worksheet.freeze_panes = "A2"

        # 自动筛选
        worksheet.auto_filter.ref = f"A1:{get_column_letter(col_count)}{row_count + 1}"


# 装饰器函数，用于在Pydantic模型中添加Excel配置
def excel_property(value: str = "", order: int = 0, converter: Optional[str] = None,
                  width: int = 20, format_pattern: str = "", ignore: bool = False):
    """
    Excel属性装饰器

    在Pydantic模型的Field中使用：
    field_name: str = Field(..., json_schema_extra=excel_property("列标题", order=1))

    Args:
        value: 列标题
        order: 列顺序
        converter: 转换器名称
        width: 列宽度
        format_pattern: 格式化模式
        ignore: 是否忽略

    Returns:
        用于json_schema_extra的配置字典
    """
    return {
        "excel": {
            "value": value,
            "order": order,
            "converter": converter,
            "width": width,
            "format_pattern": format_pattern,
            "ignore": ignore
        }
    }


# 全局导出器实例
excel_exporter = ExcelExporter()